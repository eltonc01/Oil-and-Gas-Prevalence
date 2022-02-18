import pandas as pd
from openpyxl import load_workbook
import geopy.distance
import math

df_direction = pd.read_csv('hourly wind direction.csv')
df_speed = pd.read_csv('daily wind speed.csv')
df_ozone = pd.read_csv('no2.csv')
df_wells = pd.read_excel('fracking wells.xlsx')
df_sites = pd.read_excel('prevalence sites no2.xlsx')

wb = load_workbook('template.xlsx')
sheet = wb.active


def well_activity_center(lat_input, long_input, nv, nev, ev, sev, sv, swv, wv, nwv):
    for num1 in range(0, 8):
        name = 'str'
        if num1 == 0:
            name = 'wells_north.csv'
        elif num1 == 1:
            name = 'wells_northeast.csv'
        elif num1 == 2:
            name = 'wells_east.csv'
        elif num1 == 3:
            name = 'wells_southeast.csv'
        elif num1 == 4:
            name = 'wells_south.csv'
        elif num1 == 5:
            name = 'wells_southwest.csv'
        elif num1 == 6:
            name = 'wells_west.csv'
        elif num1 == 7:
            name = 'wells_northwest.csv'
        lat = [0]
        long = [0]
        val = [0]
        toprow = {'Latitude': lat, 'Longitude': long, 'Value': val}
        df = pd.DataFrame(toprow)
        df.to_csv(name)
    loc = [lat_input, long_input]
    for row in range(0, len(df_wells)):
        temp_lat = df_wells.loc[row, 'Latitude']
        temp_long = df_wells.loc[row, 'Longitude']
        temp_loc = [temp_lat, temp_long]
        dist = geopy.distance.distance(loc, temp_loc).km
        if 20 <= dist <= 50:
            value = df_wells.loc[row, 'Oil or Gas']
            hyp = math.hypot(abs(lat_input - temp_lat), abs(long_input - temp_long))
            # print(f'Hyp: {hyp}')
            angle = 0
            adj_dist = abs(long_input - temp_long)
            adj_dist_ref = -(long_input - temp_long)
            opp = -(lat_input - temp_lat)
            # print(f'Adj: {adj_dist_ref}, Opp: {opp}')
            rad = (abs(adj_dist) / hyp)
            deg = math.degrees(math.acos(rad))
            # print(f'Deg: {deg}')
            if opp < 0:
                if adj_dist_ref < 0:
                    angle = 90 - deg + 180
                elif adj_dist_ref > 0:
                    angle = deg + 90
            elif opp > 0:
                if adj_dist_ref < 0:
                    angle = deg + 270
                elif adj_dist_ref > 0:
                    angle = 90 - deg
            # print(f'Angle: {angle}')
            if angle <= 22.5 or angle > 337.5:
                data = {
                    'Latitude': [temp_lat],
                    'Longitude': [temp_long],
                    'Value': [value]
                }
                df = pd.DataFrame(data)
                df.to_csv('wells_north.csv', mode='a', index=False, header=False)
            elif 67.5 >= angle > 22.5:
                data = {
                    'Latitude': [temp_lat],
                    'Longitude': [temp_long],
                    'Value': [value]
                }
                df = pd.DataFrame(data)
                df.to_csv('wells_northeast.csv', mode='a', index=False, header=False)
            elif 112.5 >= angle > 67.5:
                data = {
                    'Latitude': [temp_lat],
                    'Longitude': [temp_long],
                    'Value': [value]
                }
                df = pd.DataFrame(data)
                df.to_csv('wells_east.csv', mode='a', index=False, header=False)
            elif 157.5 >= angle > 112.5:
                data = {
                    'Latitude': [temp_lat],
                    'Longitude': [temp_long],
                    'Value': [value]
                }
                df = pd.DataFrame(data)
                df.to_csv('wells_southeast.csv', mode='a', index=False, header=False)
            elif 202.5 >= angle > 157.5:
                data = {
                    'Latitude': [temp_lat],
                    'Longitude': [temp_long],
                    'Value': [value]
                }
                df = pd.DataFrame(data)
                df.to_csv('wells_south.csv', mode='a', index=False, header=False)
            elif 247.5 >= angle > 202.5:
                data = {
                    'Latitude': [temp_lat],
                    'Longitude': [temp_long],
                    'Value': [value]
                }
                df = pd.DataFrame(data)
                df.to_csv('wells_southwest.csv', mode='a', index=False, header=False)
            elif 292.5 >= angle > 247.5:
                data = {
                    'Latitude': [temp_lat],
                    'Longitude': [temp_long],
                    'Value': [value]
                }
                df = pd.DataFrame(data)
                df.to_csv('wells_west.csv', mode='a', index=False, header=False)
            elif 337.5 >= angle > 292.5:
                data = {
                    'Latitude': [temp_lat],
                    'Longitude': [temp_long],
                    'Value': [value]
                }
                df = pd.DataFrame(data)
                df.to_csv('wells_northwest.csv', mode='a', index=False, header=False)
    df_north = pd.read_csv('wells_north.csv')
    df_northeast = pd.read_csv('wells_northeast.csv')
    df_east = pd.read_csv('wells_east.csv')
    df_southeast = pd.read_csv('wells_southeast.csv')
    df_south = pd.read_csv('wells_south.csv')
    df_southwest = pd.read_csv('wells_southwest.csv')
    df_west = pd.read_csv('wells_west.csv')
    df_northwest = pd.read_csv('wells_northwest.csv')
    nlat_avg = 0
    nlong_avg = 0
    for n in range(0, len(df_north)):
        nlat_avg += df_north.iloc[n, 0]
        nlong_avg += df_north.iloc[n, 1]
    if len(df_north) != 1:
        # print(nlat_avg)
        # print(nlong_avg)
        nlat_avg = nlat_avg / (len(df_north) - 1)
        nlong_avg = nlong_avg / (len(df_north) - 1)
    ncenter = [nlat_avg, nlong_avg]
    nelat_avg = 0
    nelong_avg = 0
    for n in range(0, len(df_northeast)):
        nelat_avg += df_northeast.iloc[n, 0]
        nelong_avg += df_northeast.iloc[n, 1]
    if len(df_northeast) != 1:
        nelat_avg = nelat_avg / (len(df_northeast) - 1)
        nelong_avg = nelong_avg / (len(df_northeast) - 1)
    necenter = [nelat_avg, nelong_avg]
    elat_avg = 0
    elong_avg = 0
    for n in range(0, len(df_east)):
        elat_avg += df_east.iloc[n, 0]
        elong_avg += df_east.iloc[n, 1]
    if len(df_east) != 1:
        elat_avg = elat_avg / (len(df_east) - 1)
        elong_avg = elong_avg / (len(df_east) - 1)
    ecenter = [elat_avg, elong_avg]
    selat_avg = 0
    selong_avg = 0
    for n in range(0, len(df_southeast)):
        selat_avg += df_southeast.iloc[n, 0]
        selong_avg += df_southeast.iloc[n, 1]
    if len(df_southeast) != 1:
        selat_avg = selat_avg / (len(df_southeast) - 1)
        selong_avg = selong_avg / (len(df_southeast) - 1)
    secenter = [selat_avg, selong_avg]
    slat_avg = 0
    slong_avg = 0
    for n in range(0, len(df_south)):
        slat_avg += df_south.iloc[n, 0]
        slong_avg += df_south.iloc[n, 1]
    if len(df_south) != 1:
        slat_avg = slat_avg / (len(df_south) - 1)
        slong_avg = slong_avg / (len(df_south) - 1)
    scenter = [slat_avg, slong_avg]
    swlat_avg = 0
    swlong_avg = 0
    for n in range(0, len(df_southwest)):
        swlat_avg += df_southwest.iloc[n, 0]
        swlong_avg += df_southwest.iloc[n, 1]
    if len(df_southwest) != 1:
        swlat_avg = swlat_avg / (len(df_southwest) - 1)
        swlong_avg = swlong_avg / (len(df_southwest) - 1)
    swcenter = [swlat_avg, swlong_avg]
    wlat_avg = 0
    wlong_avg = 0
    for n in range(0, len(df_west)):
        wlat_avg += df_west.iloc[n, 0]
        wlong_avg += df_west.iloc[n, 1]
    if len(df_west) != 1:
        wlat_avg = wlat_avg / (len(df_west) - 1)
        wlong_avg = wlong_avg / (len(df_west) - 1)
    wcenter = [wlat_avg, wlong_avg]
    nwlat_avg = 0
    nwlong_avg = 0
    for n in range(0, len(df_northwest)):
        nwlat_avg += df_northwest.iloc[n, 0]
        nwlong_avg += df_northwest.iloc[n, 1]
    if len(df_northwest) != 1:
        nwlat_avg = nwlat_avg / (len(df_northwest) - 1)
        nwlong_avg = nwlong_avg / (len(df_northwest) - 1)
    nwcenter = [nwlat_avg, nwlong_avg]


    return ncenter, necenter, ecenter, secenter, scenter, swcenter, wcenter, nwcenter


def wind_direction(site_lat, site_long, date):
    val = False
    n = 0
    ne = 0
    e = 0
    se = 0
    s = 0
    sw = 0
    w = 0
    nw = 0
    v = 0
    b = False
    for r in range(0, len(df_direction)):
        val = False
        if round(site_lat, 3) == round(df_direction.loc[r, 'Latitude'], 3) and round(site_long, 3) == round(
                df_direction.loc[r, 'Longitude'], 3):
            if date == df_direction.loc[r, 'Date Local']:
                val = True
                b = True
                angle = df_direction.loc[r, 'Sample Measurement']
                if angle <= 22.5 or angle > 337.5:
                    n += 1
                elif 67.5 >= angle > 22.5:
                    ne += 1
                elif 112.5 >= angle > 67.5:
                    e += 1
                elif 157.5 >= angle > 112.5:
                    se += 1
                elif 202.5 >= angle > 157.5:
                    s += 1
                elif 247.5 >= angle > 202.5:
                    sw += 1
                elif 292.5 >= angle > 247.5:
                    w += 1
                elif 337.5 >= angle > 292.5:
                    nw += 1
        if val is True:
            points = n + ne + e + se + s + sw + w + nw
            for i in range(0, 8):
                if i == 0:
                    n = n / points
                elif i == 1:
                    ne = ne / points
                elif i == 2:
                    e = e / points
                elif i == 3:
                    se = se / points
                elif i == 4:
                    s = s / points
                elif i == 5:
                    sw = sw / points
                elif i == 6:
                    w = w / points
                elif i == 7:
                    nw = nw / points
        if b is True:
            v += 1
        if v >= 25:
            break

    return n, ne, e, se, s, sw, w, nw


wb = load_workbook('template.xlsx')
sheet = wb.active

site = True
p = False
for row in range(1, len(df_ozone)):
    latitude = df_ozone.loc[row, 'Latitude']
    longitude = df_ozone.loc[row, 'Longitude']
    name = df_ozone.loc[row, 'Local Site Name']
    print(name)
    if latitude == df_ozone.loc[row - 1, 'Latitude'] and p is True:
        continue
    s = False
    t = False
    for r1 in range(0, len(df_sites)):
        if round(latitude, 4) == round(df_sites.loc[r1, 'Latitude'], 4) and round(longitude, 4) == round(
                df_sites.loc[r1, 'Longitude'], 4):
            site = r1
            if df_sites.loc[r1, 'Exempt'] == 'Yes':
                s = True
            t = True
            break
    if t is False:
        continue
    if s is True:
        continue
    date = df_ozone.loc[row, 'Date Local']
    wind = wind_direction(latitude, longitude, date)
    n = wind[0]
    ne = wind[1]
    e = wind[2]
    se = wind[3]
    s = wind[4]
    sw = wind[5]
    w = wind[6]
    nw = wind[7]
    if n == 0 and ne == 0 and e == 0 and se == 0 and s == 0 and sw == 0 and w == 0 and nw == 0:
        p = True
        continue
    p = False
    if latitude != df_ozone.loc[row - 1, 'Latitude'] or site is True:
        wells = well_activity_center(latitude, longitude, n, ne, e, se, s, sw, w, nw)
        north = wells[0]
        northeast = wells[1]
        east = wells[2]
        southeast = wells[3]
        south = wells[4]
        southwest = wells[5]
        west = wells[6]
        northwest = wells[7]
        site = False
    for r in range(0, len(df_speed)):
        if date == df_speed.loc[r, 'Date Local']:
            speed = df_speed.loc[r, 'Arithmetic Mean']
            break
    df_north1 = pd.read_csv('wells_north.csv')
    df_northeast1 = pd.read_csv('wells_northeast.csv')
    df_east1 = pd.read_csv('wells_east.csv')
    df_southeast1 = pd.read_csv('wells_southeast.csv')
    df_south1 = pd.read_csv('wells_south.csv')
    df_southwest1 = pd.read_csv('wells_southwest.csv')
    df_west1 = pd.read_csv('wells_west.csv')
    df_northwest1 = pd.read_csv('wells_northwest.csv')
    north_iter = 0
    northeast_iter = 0
    east_iter = 0
    southeast_iter = 0
    south_iter = 0
    southwest_iter = 0
    west_iter = 0
    northwest_iter = 0

    z = float(speed)
    a = 5

    for direc in range(0, 8):
        site_coord = [latitude, longitude]
        if direc == 0:
            L = geopy.distance.distance(site_coord, north).km
            s = 0.0
            for num in range(1, len(df_north1)):
                if pd.isnull(df_north1.iloc[num, 2]) is True:
                    continue
                s += float(round(df_north1.iloc[num, 2], 3))
            s = (s / 2) - ((5 * z) - 25)
            x = n
            if L != 0:
                north_iter = (15 / L) * (s) * x / a
            if s < 0:
                north_iter = 0
        elif direc == 1:
            L = geopy.distance.distance(site_coord, northeast).km
            s = 0.0
            for num in range(1, len(df_northeast1)):
                if pd.isnull(df_northeast1.iloc[num, 2]) is True:
                    continue
                s += round(float(df_northeast1.iloc[num, 2]), 3)
            s = (s / 2) - ((5 * z) - 25)
            x = ne
            if L != 0:
                northeast_iter = (15 / L) * (s) * x / a
            if s < 0:
                northeast_iter = 0
        elif direc == 2:
            L = geopy.distance.distance(site_coord, east).km
            s = 0.0
            for num in range(1, len(df_east1)):
                if pd.isnull(df_east1.iloc[num, 2]) is True:
                    continue
                s += round(float(df_east1.iloc[num, 2]), 3)
            x = e
            s = (s / 2) - ((5 * z) - 25)
            if L != 0:
                east_iter = (15 / L) * (s) * x / a
            if s < 0:
                east_iter = 0
        elif direc == 3:
            L = geopy.distance.distance(site_coord, southeast).km
            s = 0.0
            for num in range(1, len(df_southeast1)):
                if pd.isnull(df_southeast1.iloc[num, 2]) is True:
                    continue
                s += round(float(df_southeast1.iloc[num, 2]), 3)
            s = (s / 2) - ((5 * z) - 25)
            x = se
            if L != 0:
                southeast_iter = (15 / L) * (s) * x / a
            if s < 0:
                southeast_iter = 0
        elif direc == 4:
            L = geopy.distance.distance(site_coord, south).km
            s = 0.0
            for num in range(1, len(df_south1)):
                if pd.isnull(df_south1.iloc[num, 2]) is True:
                    continue
                s += round(float(df_south1.iloc[num, 2]), 3)
            s = (s / 2) - ((5 * z) - 25)
            x = s
            if L != 0:
                south_iter = (15 / L) * (s) * x / a
            if s < 0:
                south_iter = 0
        elif direc == 5:
            L = geopy.distance.distance(site_coord, southwest).km
            s = 0.0
            for num in range(1, len(df_southwest1)):
                if pd.isnull(df_southwest1.iloc[num, 2]) is True:
                    continue
                s += round(float(df_southwest1.iloc[num, 2]), 3)
            s = (s / 2) - ((5 * z) - 25)
            x = sw
            if L != 0:
                southwest_iter = (15 / L) * (s) * x / a
            if s < 0:
                southwest_iter = 0
        elif direc == 6:
            L = geopy.distance.distance(site_coord, west).km
            s = 0.0
            for num in range(1, len(df_west1)):
                if pd.isnull(df_west1.iloc[num, 2]) is True:
                    continue
                s += round(float(df_west1.iloc[num, 2]), 3)
            s = (s / 2) - ((5 * z) - 25)
            x = w
            if L != 0:
                west_iter = (15 / L) * (s) * x / a
            if s < 0:
                west_iter = 0
        elif direc == 7:
            L = geopy.distance.distance(site_coord, northwest).km
            s = 0.0
            for num in range(1, len(df_northwest1)):
                if pd.isnull(df_northwest1.iloc[num, 2]) is True:
                    continue
                s += round(float(df_northwest1.iloc[num, 2]), 3)
            s = (s / 2) - ((5 * z) - 25)
            x = nw
            if L != 0:
                northwest_iter = (15 / L) * (s) * x / a
            if s < 0:
                northwest_iter = 0

    wind_prevalence = north_iter + northeast_iter + east_iter + \
                      southeast_iter + south_iter + southwest_iter + west_iter + northwest_iter
    if wind_prevalence == 0:
        continue
    prevalence = df_sites.loc[r1, 'Prevalence']
    overall = wind_prevalence + prevalence
    print(f'{overall}: {df_ozone.loc[row, "Arithmetic Mean"]}')
    sheet.append(
        [prevalence, df_ozone.loc[row, 'Arithmetic Mean'], df_ozone.loc[row, '1st Max Value'], wind_prevalence, overall,
         name])

wb.save('wind and no2.xlsx')
